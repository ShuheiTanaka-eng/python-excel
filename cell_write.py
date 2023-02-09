import openpyxl as excel

book = excel.Workbook()
sheet = book.active

sheet["A1"] = "Hello"
sheet.cell(2,1,"Goodbye")
cell = sheet.cell(3,1)
cell.value = "Thank you"

sheet["B1"] = "こんにちは"
sheet.cell(2,2,"さようなら")
cell = sheet.cell(3,2)
cell.value = "ありがとう"

sheet["C1"] = "アニュハせよ"
sheet.cell(2,3,"アンnゆひ消せよkk")
cell = sheet.cell(3,3)
cell.value = "カムサハムニダ"

book.save("cell_write.xlsx")
