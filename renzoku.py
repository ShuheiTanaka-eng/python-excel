import openpyxl as excel

book = excel.Workbook()
sheet = book.active

for i in range(10):
    sheet.cell(i+1,1,i)
    sheet.cell(i+1,2,i*i)
    sheet.cell(i+1,3,i*i*i)
    sheet.cell(i+1,4,i*i*i*i)
    sheet.cell(i+1,5,i*i*i*i*i)
    sheet.cell(i+1,6,i*i*i*i*i*i)
    sheet.cell(i+1,7,i*i*i*i*i*i*i)
    sheet.cell(i+1,8,i*i*i*i*i*i*i*i)
    sheet.cell(i+1,9,i*i*i*i*i*i*i*i*i)
    sheet.cell(i+1,10,i*i*i*i*i*i*i*i*i*i)

book.save("renzoku.xlsx")
